# -*- coding: utf-8 -*-
from __future__ import annotations
from pathlib import Path
from typing import Dict, Iterable, List, Set
import numpy as np
import pandas as pd
from openpyxl import load_workbook
LOOKUP_PATH = Path('data/lookups/segment_assignments.csv')
MATRIX_WORKBOOK = Path('data/raw/occupation_2024_ep.xlsx')
OUTPUT_DIR = Path('data/raw/us_staffing_patterns')
NAICS_LIST_PATH = Path('data/raw/us_staffing_patterns_naics.csv')
MISSING_LOG_PATH = Path('data/raw/us_staffing_patterns_missing.csv')
NUMERIC_COLUMNS = [
    '2024 Employment',
    '2024 Percent of Industry',
    '2024 Percent of Occupation',
    'Projected 2034 Employment',
    'Projected 2034 Percent of Industry',
    'Projected 2034 Percent of Occupation',
    'Employment Change, 2024-2034',
    'Employment Percent Change, 2024-2034',
]
AGGREGATE_OVERRIDES: Dict[str, List[str]] = {
    '3270': ['3272'],
    '4840': ['4841', '4842'],
}
def get_target_naics() -> pd.Series:
    df = pd.read_csv(LOOKUP_PATH, dtype={'naics_code': str})
    codes = (df['naics_code']
             .astype(str)
             .str.strip()
             .str.replace(r'[^0-9]', '', regex=True))
    codes = codes[codes.str.len() > 0].unique()
    return pd.Series(codes).sort_values().reset_index(drop=True)
def expand_naics_codes(raw_value: object) -> Set[str]:
    if raw_value is None:
        return set()
    s = str(raw_value).strip()
    if not s or s == '?':
        return set()
    s = s.replace('\\u2013', '-').replace('\\u2014', '-').replace('\\uFFFD', '')
    if '(' in s:
        s = s.split('(')[0].strip()
    codes: Set[str] = set()
    if ',' in s:
        parts = [part.strip() for part in s.split(',') if part.strip()]
        if parts:
            base_digits = ''.join(ch for ch in parts[0] if ch.isdigit())
            if len(base_digits) >= 4:
                codes.add(base_digits[:4])
            prefix = base_digits[:-1] if len(base_digits) > 0 else ''
            for part in parts[1:]:
                digits = ''.join(ch for ch in part if ch.isdigit())
                if len(digits) >= 4:
                    codes.add(digits[:4])
                elif prefix and digits:
                    candidate = (prefix + digits)[-4:]
                    if len(candidate) == 4:
                        codes.add(candidate)
        return codes
    if '-' in s:
        segments = [segment.strip() for segment in s.split('-') if segment.strip()]
        for segment in segments:
            digits = ''.join(ch for ch in segment if ch.isdigit())
            if len(digits) >= 4:
                codes.add(digits[:4])
        if codes:
            return codes
    digits = ''.join(ch for ch in s if ch.isdigit())
    if len(digits) >= 4:
        codes.add(digits[:4])
    return codes
def build_hyperlink_map(target_codes: Iterable[str]) -> Dict[str, str]:
    wb = load_workbook(MATRIX_WORKBOOK, read_only=False, data_only=True)
    ws = wb['Table 1.9']
    desired = set(str(code) for code in target_codes)
    mapping: Dict[str, tuple[int, str]] = {}
    for row in ws.iter_rows(min_row=3):
        industry_type = (row[2].value or '').strip().lower()
        hyperlink = row[4].hyperlink.target if row[4].hyperlink else None
        if not hyperlink:
            continue
        codes = expand_naics_codes(row[3].value)
        if not codes:
            continue
        priority = 0 if industry_type == 'summary' else 1
        for code in codes:
            if code in desired and (code not in mapping or priority < mapping[code][0]):
                mapping[code] = (priority, hyperlink)
            elif code in AGGREGATE_OVERRIDES:
                for override in AGGREGATE_OVERRIDES[code]:
                    if override in desired and (override not in mapping or priority < mapping[override][0]):
                        mapping[override] = (priority + 1, hyperlink)
    wb.close()
    return {code: link for code, (priority, link) in mapping.items()}
def fetch_bls_table(code: str, url: str) -> pd.DataFrame:
    tables = pd.read_html(url, header=0)
    if not tables:
        raise ValueError(f'No tables found for {code}')
    df = tables[0]
    df = df[df['Occupation Title'] != 'Filter by Title:'].copy()
    if df.empty:
        raise ValueError(f'Empty table for {code}')
    df.insert(0, 'naics_code', code)
    df['source_url'] = url
    for col in NUMERIC_COLUMNS:
        if col in df.columns:
            df[col] = (df[col]
                       .astype(str)
                       .str.replace(',', '', regex=False)
                       .str.replace('%', '', regex=False))
            df[col] = pd.to_numeric(df[col], errors='coerce')
    return df
def main() -> None:
    codes = get_target_naics()
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    codes.to_csv(NAICS_LIST_PATH, index=False, header=['naics_code'])
    link_map = build_hyperlink_map(codes)
    missing_links: List[str] = []
    failures: List[Dict[str, str]] = []
    for code in codes:
        code_str = str(code)
        url = link_map.get(code_str)
        if not url:
            missing_links.append(code_str)
            continue
        try:
            df = fetch_bls_table(code_str, url)
        except Exception as exc:  # noqa: BLE001
            failures.append({'naics_code': code_str, 'source_url': url, 'error': str(exc)})
            continue
        out_path = OUTPUT_DIR / f'us_staffing_{code_str}.csv'
        df.to_csv(out_path, index=False)
    if missing_links or failures:
        records: List[Dict[str, str]] = []
        records.extend({'naics_code': code, 'issue': 'missing_link'} for code in missing_links)
        records.extend(failures)
        pd.DataFrame(records).to_csv(MISSING_LOG_PATH, index=False)
    elif MISSING_LOG_PATH.exists():
        MISSING_LOG_PATH.unlink()
if __name__ == '__main__':
    main()
